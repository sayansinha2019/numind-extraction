"""Tests for JSON-to-Excel conversion across document extraction scenarios."""

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from excel_export import (
    SINGLE_SHEET_NAME,
    _flatten_fields,
    _format_cell_value,
    _is_table,
    _sanitize_sheet_name,
    write_excel_from_extraction,
)
from extract import deep_merge_data, merge_strings


def _trim_row(row: tuple[object, ...]) -> list[object]:
    values = list(row)
    while values and values[-1] is None:
        values.pop()
    return values


def sheet_table(ws) -> list[list[object]]:
    """Read a worksheet as a list of rows."""
    return [_trim_row(row) for row in ws.iter_rows(values_only=True)]


def _section_start(rows: list[list[object]], section_name: str) -> int:
    for index, row in enumerate(rows):
        if row and row[0] == section_name:
            return index
    raise AssertionError(f"Section {section_name!r} not found")


def section_to_dict(ws, section_name: str = "Summary") -> dict[str, str]:
    """Read a Field/Value section into a dictionary."""
    rows = sheet_table(ws)
    start = _section_start(rows, section_name)
    header = rows[start + 1]
    if header[:2] != ["Field", "Value"]:
        raise AssertionError(f"Expected Field/Value headers in {section_name!r}")
    result: dict[str, str] = {}
    for row in rows[start + 2 :]:
        if not row or row[0] is None:
            break
        if row[0] in {"Summary", "Pages", "Errors"} or (
            len(row) > 1 and row[1] is None and row[0] not in result
        ):
            break
        result[row[0]] = row[1] if len(row) > 1 else ""
    return result


def section_table(ws, section_name: str) -> list[list[object]]:
    """Read a table section (title row, header row, then data rows)."""
    rows = sheet_table(ws)
    start = _section_start(rows, section_name)
    table_rows: list[list[object]] = []
    for row in rows[start + 1 :]:
        if not row or row[0] is None:
            break
        if row[0] in {"Summary", "Pages", "Errors"}:
            break
        table_rows.append(row)
    return table_rows


def pages_section(ws) -> list[list[object]]:
    """Read the Pages section including its header row."""
    rows = sheet_table(ws)
    start = _section_start(rows, "Pages")
    section_rows: list[list[object]] = []
    for row in rows[start + 1 :]:
        if not row or row[0] is None:
            break
        if row[0] == "Errors":
            break
        section_rows.append(row)
    return section_rows


def errors_section(ws) -> list[list[object]]:
    """Read the Errors section including its header row."""
    rows = sheet_table(ws)
    start = _section_start(rows, "Errors")
    return [row for row in rows[start + 1 :] if row and row[0] is not None]


class TestExcelHelpers(unittest.TestCase):
    def test_format_cell_value_types(self) -> None:
        self.assertEqual(_format_cell_value(None), "")
        self.assertEqual(_format_cell_value(True), "true")
        self.assertEqual(_format_cell_value(False), "false")
        self.assertEqual(_format_cell_value(42), "42")
        self.assertEqual(_format_cell_value(3.5), "3.5")
        self.assertEqual(_format_cell_value("hello"), "hello")
        self.assertEqual(_format_cell_value(["a", "b"]), "a, b")

    def test_is_table(self) -> None:
        self.assertTrue(_is_table([{"a": 1}, {"a": 2}]))
        self.assertFalse(_is_table([]))
        self.assertFalse(_is_table(["a", "b"]))
        self.assertFalse(_is_table([{"a": 1}, "bad"]))

    def test_flatten_nested_fields(self) -> None:
        rows = _flatten_fields(
            {
                "name": "John",
                "address": {"city": "NYC", "zip": "10001"},
                "tags": ["vip", "new"],
            }
        )
        self.assertEqual(
            dict(rows),
            {
                "name": "John",
                "address.city": "NYC",
                "address.zip": "10001",
                "tags": "vip, new",
            },
        )

    def test_flatten_skips_table_arrays(self) -> None:
        rows = _flatten_fields(
            {
                "invoice_id": "INV-1",
                "line_items": [{"item": "pen", "qty": 2}],
            }
        )
        self.assertEqual(dict(rows), {"invoice_id": "INV-1"})
        self.assertNotIn("line_items", dict(rows))

    def test_sanitize_sheet_name(self) -> None:
        used: set[str] = set()
        self.assertEqual(_sanitize_sheet_name("line_items", used), "line_items")
        self.assertEqual(_sanitize_sheet_name("line/items", used), "line_items_1")
        self.assertEqual(_sanitize_sheet_name("line_items", used), "line_items_2")


class TestExcelExportScenarios(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.output_dir = Path(self.temp_dir.name)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def _export(self, extracted_data: dict, name: str = "doc.pdf") -> Path:
        excel_path = self.output_dir / "output.xlsx"
        write_excel_from_extraction(extracted_data, excel_path, source_name=name)
        return excel_path

    def _workbook(self, excel_path: Path):
        return load_workbook(excel_path, read_only=True, data_only=True)

    def _sheet(self, excel_path: Path):
        wb = self._workbook(excel_path)
        self.assertEqual(wb.sheetnames, [SINGLE_SHEET_NAME])
        return wb, wb[SINGLE_SHEET_NAME]

    def test_simple_flat_document(self) -> None:
        excel_path = self._export(
            {
                "page_count": 1,
                "data": {"invoice_number": "INV-100", "total": 250.5},
                "pages": [
                    {
                        "page": 1,
                        "status": "ok",
                        "data": {"invoice_number": "INV-100", "total": 250.5},
                    }
                ],
                "errors": [],
            }
        )
        wb, sheet = self._sheet(excel_path)
        summary = section_to_dict(sheet)
        self.assertEqual(summary["source_file"], "doc.pdf")
        self.assertEqual(summary["page_count"], "1")
        self.assertEqual(summary["invoice_number"], "INV-100")
        self.assertEqual(summary["total"], "250.5")
        wb.close()

    def test_nested_fields_and_cross_page_merge(self) -> None:
        excel_path = self._export(
            {
                "page_count": 2,
                "data": {
                    "patient_name": "John Doe",
                    "address": "123 Main Street, NYC",
                    "policy": {"provider": "Acme", "id": "P-55"},
                },
                "pages": [
                    {
                        "page": 1,
                        "status": "ok",
                        "data": {
                            "patient_name": "John Doe",
                            "address": "123 Main",
                            "policy": {"provider": "Acme"},
                        },
                    },
                    {
                        "page": 2,
                        "status": "ok",
                        "data": {
                            "address": "Street, NYC",
                            "policy": {"id": "P-55"},
                        },
                    },
                ],
                "errors": [],
            }
        )
        wb, sheet = self._sheet(excel_path)
        summary = section_to_dict(sheet)
        self.assertEqual(summary["patient_name"], "John Doe")
        self.assertEqual(summary["address"], "123 Main Street, NYC")
        self.assertEqual(summary["policy.provider"], "Acme")
        self.assertEqual(summary["policy.id"], "P-55")

        pages = pages_section(sheet)
        self.assertEqual(pages[0], ["Page", "Field", "Value"])
        self.assertIn(["1", "address", "123 Main"], pages)
        self.assertIn(["2", "policy.id", "P-55"], pages)
        wb.close()

    def test_top_level_table_section(self) -> None:
        excel_path = self._export(
            {
                "page_count": 1,
                "data": {
                    "invoice_number": "INV-200",
                    "line_items": [
                        {"description": "Widget", "qty": 2, "price": 10},
                        {"description": "Gadget", "qty": 1, "price": 25},
                    ],
                },
                "pages": [{"page": 1, "status": "ok", "data": {}}],
                "errors": [],
            }
        )
        wb, sheet = self._sheet(excel_path)
        table = section_table(sheet, "line_items")
        self.assertEqual(table[0], ["description", "qty", "price"])
        self.assertEqual(table[1], ["Widget", "2", "10"])
        self.assertEqual(table[2], ["Gadget", "1", "25"])
        wb.close()

    def test_nested_table_section(self) -> None:
        excel_path = self._export(
            {
                "page_count": 1,
                "data": {
                    "invoice": {
                        "number": "INV-300",
                        "line_items": [
                            {"sku": "A1", "amount": 100},
                            {"sku": "B2", "amount": 50},
                        ],
                    }
                },
                "pages": [{"page": 1, "status": "ok", "data": {}}],
                "errors": [],
            }
        )
        wb, sheet = self._sheet(excel_path)
        summary = section_to_dict(sheet)
        self.assertEqual(summary["invoice.number"], "INV-300")

        table = section_table(sheet, "invoice.line_items")
        self.assertEqual(table[0], ["sku", "amount"])
        self.assertEqual(table[1], ["A1", "100"])
        wb.close()

    def test_multiple_table_sections(self) -> None:
        excel_path = self._export(
            {
                "page_count": 1,
                "data": {
                    "items": [{"name": "A"}],
                    "payments": [{"method": "card", "amount": 10}],
                },
                "pages": [],
                "errors": [],
            }
        )
        wb, sheet = self._sheet(excel_path)
        self.assertEqual(section_table(sheet, "items")[1], ["A"])
        self.assertEqual(section_table(sheet, "payments")[1], ["card", "10"])
        wb.close()

    def test_failed_page_and_errors_section(self) -> None:
        excel_path = self._export(
            {
                "page_count": 2,
                "data": {"field_a": "value"},
                "pages": [
                    {"page": 1, "status": "ok", "data": {"field_a": "value"}},
                    {
                        "page": 2,
                        "status": "error",
                        "error": "GPU timeout",
                        "data": {},
                    },
                ],
                "errors": [
                    {"stage": "extraction", "page": 2, "error": "GPU timeout"},
                    {"stage": "template", "pages": [1], "error": "partial template issue"},
                ],
            }
        )
        wb, sheet = self._sheet(excel_path)

        pages = pages_section(sheet)
        self.assertIn(["2", "status", "error"], pages)
        self.assertIn(["2", "error", "GPU timeout"], pages)

        errors = errors_section(sheet)
        self.assertEqual(errors[0], ["Stage", "Page(s)", "Error"])
        self.assertEqual(errors[1], ["extraction", "2", "GPU timeout"])
        self.assertEqual(errors[2], ["template", "1", "partial template issue"])
        wb.close()

    def test_empty_document_data(self) -> None:
        excel_path = self._export(
            {
                "page_count": 0,
                "data": {},
                "pages": [],
                "errors": [],
            }
        )
        wb, sheet = self._sheet(excel_path)
        summary = section_to_dict(sheet)
        self.assertEqual(summary["source_file"], "doc.pdf")
        self.assertEqual(summary["page_count"], "0")

        rows = sheet_table(sheet)
        self.assertNotIn("Errors", [row[0] for row in rows if row])
        wb.close()

    def test_boolean_and_null_fields(self) -> None:
        excel_path = self._export(
            {
                "page_count": 1,
                "data": {
                    "active": True,
                    "archived": False,
                    "notes": None,
                },
                "pages": [],
                "errors": [],
            }
        )
        wb, sheet = self._sheet(excel_path)
        summary = section_to_dict(sheet)
        self.assertEqual(summary["active"], "true")
        self.assertEqual(summary["archived"], "false")
        self.assertIn(summary.get("notes"), ("", None))
        wb.close()

    def test_merge_then_excel_end_to_end(self) -> None:
        """Simulate multi-page extraction merge and verify Excel output."""
        page_one = {
            "vendor": "Acme Corp",
            "address": "123 Main",
            "line_items": [{"item": "Paper", "qty": 5}],
        }
        page_two = {
            "address": "Street, Boston",
            "line_items": [{"item": "Ink", "qty": 2}],
            "total": 99.5,
        }

        merged = deep_merge_data(deep_merge_data({}, page_one), page_two)
        self.assertEqual(merged["address"], merge_strings("123 Main", "Street, Boston"))

        excel_path = self._export(
            {
                "page_count": 2,
                "data": merged,
                "pages": [
                    {"page": 1, "status": "ok", "data": page_one},
                    {"page": 2, "status": "ok", "data": page_two},
                ],
                "errors": [],
            },
            name="invoice.pdf",
        )

        wb, sheet = self._sheet(excel_path)
        summary = section_to_dict(sheet)
        self.assertEqual(summary["vendor"], "Acme Corp")
        self.assertEqual(summary["address"], "123 Main Street, Boston")
        self.assertEqual(summary["total"], "99.5")

        table = section_table(sheet, "line_items")
        self.assertEqual(len(table), 3)  # header + 2 rows
        self.assertEqual(table[1][0], "Paper")
        self.assertEqual(table[2][0], "Ink")
        wb.close()

    def test_table_cells_with_nested_values(self) -> None:
        excel_path = self._export(
            {
                "page_count": 1,
                "data": {
                    "rows": [
                        {"name": "A", "meta": {"color": "red"}},
                        {"name": "B", "meta": {"color": "blue"}},
                    ]
                },
                "pages": [],
                "errors": [],
            }
        )
        wb, sheet = self._sheet(excel_path)
        table = section_table(sheet, "rows")
        self.assertEqual(table[0], ["name", "meta"])
        self.assertEqual(table[1][1], "{'color': 'red'}")
        wb.close()


if __name__ == "__main__":
    unittest.main()
