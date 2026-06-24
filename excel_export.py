"""Convert extracted document JSON into a structured Excel workbook."""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

INVALID_SHEET_CHARS = re.compile(r"[\[\]\*\?:/\\]")
MAX_SHEET_NAME_LEN = 31


def _sanitize_sheet_name(name: str, used_names: set[str]) -> str:
    """Return a valid, unique Excel sheet name."""
    cleaned = INVALID_SHEET_CHARS.sub("_", name.strip()) or "Sheet"
    cleaned = cleaned[:MAX_SHEET_NAME_LEN]

    candidate = cleaned
    index = 1
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{cleaned[: MAX_SHEET_NAME_LEN - len(suffix)]}{suffix}"
        index += 1

    used_names.add(candidate)
    return candidate


def _format_cell_value(value: object) -> str:
    """Convert JSON values into Excel-friendly text."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        if not value:
            return ""
        if all(not isinstance(item, (dict, list)) for item in value):
            return ", ".join(_format_cell_value(item) for item in value)
        return ""
    return str(value)


def _is_table(value: object) -> bool:
    """Return True when a value should be exported as a table sheet."""
    return isinstance(value, list) and bool(value) and all(isinstance(item, dict) for item in value)


def _flatten_fields(
    data: dict,
    *,
    prefix: str = "",
) -> list[tuple[str, str]]:
    """Flatten nested dictionaries into dotted field paths."""
    rows: list[tuple[str, str]] = []

    for key, value in data.items():
        field_name = f"{prefix}.{key}" if prefix else str(key)

        if isinstance(value, dict):
            rows.extend(_flatten_fields(value, prefix=field_name))
        elif _is_table(value):
            continue
        else:
            rows.append((field_name, _format_cell_value(value)))

    return rows


def _iter_tables(data: dict, *, prefix: str = "") -> list[tuple[str, list[dict]]]:
    """Find list-of-dict tables at any nesting level."""
    tables: list[tuple[str, list[dict]]] = []

    for key, value in data.items():
        field_name = f"{prefix}.{key}" if prefix else str(key)

        if _is_table(value):
            tables.append((field_name, value))
        elif isinstance(value, dict):
            tables.extend(_iter_tables(value, prefix=field_name))

    return tables


def _collect_table_columns(rows: list[dict]) -> list[str]:
    """Collect ordered unique columns from a list of row dictionaries."""
    columns: list[str] = []

    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)

    return columns


def _autosize_columns(worksheet) -> None:
    """Set a reasonable column width based on content."""
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _write_header_row(worksheet, headers: list[str]) -> None:
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)


def _write_key_value_sheet(worksheet, rows: list[tuple[str, str]]) -> None:
    _write_header_row(worksheet, ["Field", "Value"])
    for field, value in rows:
        worksheet.append([field, value if value is not None else ""])
    _autosize_columns(worksheet)


def _write_table_sheet(worksheet, rows: list[dict]) -> None:
    columns = _collect_table_columns(rows)
    _write_header_row(worksheet, columns)
    for row in rows:
        worksheet.append([_format_cell_value(row.get(column, "")) for column in columns])
    _autosize_columns(worksheet)


def _page_label(page_entry: dict) -> str:
    """Format one or more page numbers for Excel output."""
    pages = page_entry.get("pages")
    if pages:
        return ", ".join(str(page) for page in pages)
    if page_entry.get("page") is not None:
        return str(page_entry["page"])
    return ""


def write_excel_from_extraction(
    extracted_data: dict,
    excel_path: Path,
    *,
    source_name: str,
) -> None:
    """Write the final merged extraction result to an Excel workbook."""
    workbook = Workbook()
    used_sheet_names: set[str] = set()

    summary_sheet = workbook.active
    summary_sheet.title = _sanitize_sheet_name("Summary", used_sheet_names)

    summary_rows = [
        ("source_file", source_name),
        ("page_count", _format_cell_value(extracted_data.get("page_count", ""))),
    ]
    summary_rows.extend(_flatten_fields(extracted_data.get("data", {})))
    _write_key_value_sheet(summary_sheet, summary_rows)

    document_data = extracted_data.get("data", {})
    for table_name, table_rows in _iter_tables(document_data):
        sheet_name = _sanitize_sheet_name(table_name, used_sheet_names)
        table_sheet = workbook.create_sheet(title=sheet_name)
        _write_table_sheet(table_sheet, table_rows)

    pages_sheet = workbook.create_sheet(
        title=_sanitize_sheet_name("Pages", used_sheet_names)
    )
    _write_header_row(pages_sheet, ["Page", "Field", "Value"])
    for page_entry in extracted_data.get("pages", []):
        page_label = _page_label(page_entry)
        status = page_entry.get("status", "ok")
        if status != "ok":
            pages_sheet.append([page_label, "status", status])
            if page_entry.get("error"):
                pages_sheet.append([page_label, "error", page_entry["error"]])
            continue

        for field, value in _flatten_fields(page_entry.get("data", {})):
            pages_sheet.append([page_label, field, value])
    _autosize_columns(pages_sheet)

    errors = extracted_data.get("errors", [])
    if errors:
        errors_sheet = workbook.create_sheet(
            title=_sanitize_sheet_name("Errors", used_sheet_names)
        )
        _write_header_row(errors_sheet, ["Stage", "Page(s)", "Error"])
        for error in errors:
            pages = error.get("pages")
            if pages is None and error.get("page") is not None:
                pages = [error["page"]]
            page_label = ", ".join(str(page) for page in pages) if pages else ""
            errors_sheet.append([error.get("stage", ""), page_label, error.get("error", "")])
        _autosize_columns(errors_sheet)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)
